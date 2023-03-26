from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as W
from selenium.webdriver.support import expected_conditions as E
from selenium.webdriver.support.ui import Select
import time
import openpyxl as O

exec_path = r"E:\trud\geckodriver.exe"
URL = "http://172.1.4.158"
Excel_file = "E:\\trud\\test.xlsx"
Exel_worksheet = "Sheet1"
wait_time_out = 10
link_css_locator = ".nav-tabs > li:nth-child(3) > a"
bn_Add_locator = ".ui-grid-pager-count > .btn"
id_Element_locator = "insId"
name_Element_locator = "emName"
work_Element_locator = "erEat"
work_att_Elemen_locator = "erAtt"
data_Element_locator = "eeActionDt"
text_Element_locator = "eeText"
type_Element_locator = "eeType"
data_type_Element_locator = "eeDt"
number_Element_locator = "eeNum"
bn_locator = ".tk-modal-footer > .btn:nth-child(2)"
bn_end_locator = ".btn-warning"
username_locator = "loginItem_Login"
password_locator = "loginItem_Pwd"
enter_locator = ".btn"
shurnal_locator = ".col-lg-3:nth-child(1) .panel-footer .fa"
shurnal1_locator = ".dropdown:nth-child(2) > .dropdown-toggle"
shurnal2_locator = "#\31 626631788716-5-uiGrid-000J-cell .fa"
vhid_locator = ".fa-file-text-o"
driver = webdriver.Firefox(executable_path = exec_path)
driver.get(URL)
wait_variable = W(driver, wait_time_out)
driver.execute_script("window.scrollBy(0,240)","")
user_elm = wait_variable.until(E.presence_of_element_located((By.ID,username_locator)))
pass_elm = wait_variable.until(E.presence_of_element_located((By.ID,password_locator)))
enter = wait_variable.until(E.presence_of_element_located((By.CSS_SELECTOR,enter_locator)))
user = str("u24001-0003")
user_elm.send_keys(user)
passw = str("xZ426NMvnU")
pass_elm.send_keys(passw)
enter.click()
shurnal1 = wait_variable.until(E.presence_of_element_located((By.CSS_SELECTOR,shurnal1_locator)))
shurnal1.click()
shurnal2 = wait_variable.until(E.presence_of_element_located((By.CSS_SELECTOR,shurnal2_locator)))
shurnal2.click()
time.sleep(1)
vhid = wait_variable.until(E.presence_of_element_located((By.CSS_SELECTOR,vhid_locator)))
vhid.click()
time.sleep(1)
linktex = wait_variable.until(E.presence_of_element_located((By.CSS_SELECTOR,link_css_locator)))
linktex.click()
time.sleep(1)
bnAdd = wait_variable.until(E.presence_of_element_located((By.CSS_SELECTOR,bn_Add_locator)))
time.sleep(1)
bnAdd.click()
time.sleep(1)
id_Element =  wait_variable.until(E.presence_of_element_located((By.ID,id_Element_locator)))
name_Element =  wait_variable.until(E.presence_of_element_located((By.ID,name_Element_locator)))
work_Element =  Select(wait_variable.until(E.presence_of_element_located((By.ID,work_Element_locator))))
work_att_Elemen =  Select(wait_variable.until(E.presence_of_element_located((By.ID,work_att_Elemen_locator))))
data_Element =  wait_variable.until(E.presence_of_element_located((By.ID,data_Element_locator)))
text_Element =  wait_variable.until(E.presence_of_element_located((By.ID,text_Element_locator)))
type_Element =  wait_variable.until(E.presence_of_element_located((By.ID,type_Element_locator)))
data_type_Element =  wait_variable.until(E.presence_of_element_located((By.ID,data_type_Element_locator)))
number_Element =  wait_variable.until(E.presence_of_element_located((By.ID,number_Element_locator)))
bn =  wait_variable.until(E.presence_of_element_located((By.CSS_SELECTOR,bn_locator)))
bn_end =  wait_variable.until(E.presence_of_element_located((By.CSS_SELECTOR,bn_end_locator)))
wb = O.load_workbook(Excel_file)
ws = wb[Exel_worksheet]
for r in range(2, ws.max_row + 1):
    a = str(ws.cell(r,1).value)
    id_Element.clear()
    id_Element.send_keys(a)
    b = str(ws.cell(r,2).value)
    name_Element.send_keys(b)
    c = str(ws.cell(r,3).value)
    work_Element.select_by_visible_text(c.strip())
    d = str(ws.cell(r,4).value)
    time.sleep(1)
    work_att_Elemen.select_by_visible_text(d.strip())
    e = str(ws.cell(r,5).value)
    data_Element.send_keys(e.strip())
    #time.sleep(5)
    f = str(ws.cell(r,6).value)
    text_Element.send_keys(f)
    g = str(ws.cell(r,7).value)
    type_Element.send_keys(g)
    j = str(ws.cell(r,8).value)
    data_type_Element.send_keys(j.strip())
    i = str(ws.cell(r,9).value)
    number_Element.send_keys(i)
    bn.click()
    time.sleep(2)
    ws.cell(r,10).value = "Pass"
wb.save(Excel_file)
wb.close()
bn_end.click()
